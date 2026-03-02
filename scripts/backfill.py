"""
Parse timesheet xlsx files and generate SQL to backfill historical data.
Generates: supabase/backfill.sql
"""
import openpyxl
from datetime import date, timedelta, datetime
import os

PROJECT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ── 1. Read all xlsx files ──────────────────────────────────────────
FILES = [
    "Timesheets - Mar 1 - Jun 1, 2025.xlsx",
    "Timesheets - Jun 1 - Sep 1, 2025.xlsx",
    "Timesheets - Dec 1 - Mar 1, 2026.xlsx",
    # Sep 1 - Sep 1 2026 is empty, skip
]

entries = []  # list of dicts

for fname in FILES:
    path = os.path.join(PROJECT, fname)
    wb = openpyxl.load_workbook(path, data_only=True)
    # Use first sheet (Makelab, Inc or Entries)
    ws = wb[wb.sheetnames[0]]

    # Read header row
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        first = (data.get("First Name") or "").strip()
        last = (data.get("Last Name") or "").strip()
        if not first or not last:
            continue

        full_name = f"{first} {last}"
        dt = data.get("Date")
        start = data.get("Start Time")
        end = data.get("End Time")
        breaks = data.get("Unpaid Breaks") or 0
        regular = data.get("Regular") or 0
        rate = data.get("Hourly Rate") or 0

        if not dt or not start or not end:
            continue

        # Parse datetime values
        if isinstance(dt, datetime):
            work_date = dt.date()
        elif isinstance(dt, str):
            work_date = datetime.fromisoformat(dt).date()
        else:
            continue

        if isinstance(start, datetime):
            start_time = start.strftime("%H:%M")
        elif isinstance(start, str):
            start_time = datetime.fromisoformat(start).strftime("%H:%M")
        else:
            continue

        if isinstance(end, datetime):
            end_time = end.strftime("%H:%M")
        elif isinstance(end, str):
            end_time = datetime.fromisoformat(end).strftime("%H:%M")
        else:
            continue

        break_min = int(float(breaks)) if breaks else 0
        net_hours = round(float(regular), 2) if regular else 0

        # Calculate gross hours from start/end
        sh, sm = map(int, start_time.split(":"))
        eh, em = map(int, end_time.split(":"))
        gross_hours = round((eh * 60 + em - sh * 60 - sm) / 60.0, 2)

        entries.append({
            "full_name": full_name,
            "work_date": work_date,
            "start_time": start_time,
            "end_time": end_time,
            "break_minutes": break_min,
            "gross_hours": gross_hours,
            "net_hours": net_hours,
        })

print(f"Parsed {len(entries)} time entries")

# Filter out employees to exclude
EXCLUDE = {"Jiwon Lee", "Nicholas Hongxi Yuan", "Olga Kassian"}
entries = [e for e in entries if e["full_name"] not in EXCLUDE]
print(f"After excluding {EXCLUDE}: {len(entries)} entries")

# ── 2. Determine unique employees ──────────────────────────────────
employee_names = sorted(set(e["full_name"] for e in entries))
print(f"Employees: {employee_names}")

# ── 3. Determine date range and generate pay periods ────────────────
# Existing periods start Feb 7 2026, biweekly (14 days each).
# Work backwards to cover all historical data.
all_dates = [e["work_date"] for e in entries]
min_date = min(all_dates)
max_date = max(all_dates)
print(f"Date range: {min_date} to {max_date}")

# Anchor: Feb 7, 2026. Go backwards by 14-day increments.
anchor = date(2026, 2, 7)
periods = []

# Go backwards until we pass min_date
period_start = anchor
while period_start > min_date - timedelta(days=14):
    period_start -= timedelta(days=14)

# Now generate forward until we pass max_date
current = period_start
while current <= max_date:
    period_end = current + timedelta(days=13)
    periods.append((current, period_end))
    current += timedelta(days=14)

# Split into new periods (to create) and existing periods (already in DB)
new_periods = [(s, e) for s, e in periods if s < anchor]
existing_periods = [(s, e) for s, e in periods if s >= anchor]
print(f"Historical pay periods to create: {len(new_periods)}")

# Also include existing DB periods for entries that fall within them
# Known existing periods from DB: Feb 7-20, Feb 21-Mar 6
db_periods = [
    (date(2026, 2, 7), date(2026, 2, 20)),
    (date(2026, 2, 21), date(2026, 3, 6)),
]
all_periods = new_periods + db_periods

# ── 4. Assign entries to periods ────────────────────────────────────
def find_period(work_date):
    for s, e in all_periods:
        if s <= work_date <= e:
            return (s, e)
    return None

orphans = []
for entry in entries:
    pp = find_period(entry["work_date"])
    if pp:
        entry["period_start"] = pp[0]
        entry["period_end"] = pp[1]
    else:
        orphans.append(entry)

if orphans:
    print(f"WARNING: {len(orphans)} entries don't fit any period:")
    for o in orphans:
        print(f"  {o['full_name']} {o['work_date']}")

# Filter to only entries with assigned periods
entries = [e for e in entries if "period_start" in e]

# ── 5. Generate SQL ─────────────────────────────────────────────────
sql_lines = []
sql_lines.append("-- ============================================================")
sql_lines.append("-- Backfill historical timesheet data")
sql_lines.append(f"-- Generated from {len(FILES)} xlsx files")
sql_lines.append(f"-- {len(entries)} time entries for {len(employee_names)} employees")
sql_lines.append("-- ============================================================")
sql_lines.append("")
sql_lines.append("-- Temporarily disable triggers that would interfere")
sql_lines.append("ALTER TABLE time_entries DISABLE TRIGGER trg_compute_hours;")
sql_lines.append("ALTER TABLE time_entries DISABLE TRIGGER trg_validate_work_date;")
sql_lines.append("ALTER TABLE time_entries DISABLE TRIGGER trg_prevent_locked_edits;")
sql_lines.append("")

# Insert pay periods (only new historical ones, not existing DB ones)
sql_lines.append("-- ── Pay Periods ──────────────────────────────────────────────")
for s, e in sorted(new_periods):
    sql_lines.append(
        f"INSERT INTO pay_periods (start_date, end_date, status) "
        f"VALUES ('{s}', '{e}', 'closed') "
        f"ON CONFLICT (start_date) DO NOTHING;"
    )
sql_lines.append("")

# Insert time entries grouped by employee
sql_lines.append("-- ── Time Entries ────────────────────────────────────────────")
sql_lines.append("-- Each entry uses a subquery to look up employee_id and pay_period_id")
sql_lines.append("")

for entry in sorted(entries, key=lambda e: (e["full_name"], e["work_date"])):
    name = entry["full_name"].replace("'", "''")
    sql_lines.append(
        f"INSERT INTO time_entries (employee_id, pay_period_id, work_date, "
        f"start_time, end_time, break_minutes, gross_hours, net_hours) "
        f"VALUES ("
        f"(SELECT id FROM employees WHERE full_name = '{name}'), "
        f"(SELECT id FROM pay_periods WHERE start_date = '{entry['period_start']}'), "
        f"'{entry['work_date']}', "
        f"'{entry['start_time']}', '{entry['end_time']}', "
        f"{entry['break_minutes']}, {entry['gross_hours']}, {entry['net_hours']}"
        f") ON CONFLICT (employee_id, work_date) DO NOTHING;"
    )

sql_lines.append("")
sql_lines.append("-- Re-enable triggers")
sql_lines.append("ALTER TABLE time_entries ENABLE TRIGGER trg_compute_hours;")
sql_lines.append("ALTER TABLE time_entries ENABLE TRIGGER trg_validate_work_date;")
sql_lines.append("ALTER TABLE time_entries ENABLE TRIGGER trg_prevent_locked_edits;")
sql_lines.append("")
sql_lines.append("-- Done! Verify with:")
sql_lines.append("-- SELECT e.full_name, count(*) as entries, sum(net_hours) as total_hours")
sql_lines.append("-- FROM time_entries t JOIN employees e ON e.id = t.employee_id")
sql_lines.append("-- GROUP BY e.full_name ORDER BY e.full_name;")

# Write SQL file
out_path = os.path.join(PROJECT, "supabase", "backfill.sql")
with open(out_path, "w") as f:
    f.write("\n".join(sql_lines) + "\n")

print(f"\nGenerated: {out_path}")
print(f"Run this SQL in Supabase SQL Editor to backfill the data.")
print(f"\nIMPORTANT: Employee names in the SQL must match full_name in your employees table:")
for name in employee_names:
    print(f"  - {name}")
print(f"\nIf any names don't match, update them in the employees table first.")
